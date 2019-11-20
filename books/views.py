import json

from django.contrib import messages
from django.core import serializers
from django.core.paginator import Paginator
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from django.shortcuts import render, render_to_response

# Create your views here.
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from django.views.generic import ListView, FormView, DetailView

from books.forms import CommentForm
from books.models import Book
from openpyxl import Workbook
from openpyxl.styles import Border, Font


class BookList(ListView):
    model = Book
    context_object_name = 'books'
    template_name = 'books/list.html'
    paginate_by = 10

class BookDetail(DetailView, FormView):
    model = Book
    template_name = 'books/details.html'
    form_class = CommentForm



def books_list(request):
    books = Book.objects.all()
    page = request.GET.get('page')
    paginator = Paginator(books, 10)
    books = paginator.get_page(page)

    format = request.GET.get('format')
    if format == "json":
        books = serializers.serialize('json', books)
        books = json.loads(books)
        return JsonResponse(books, safe=False)

    return render(
        request,
        "books/list.html",
        context={'books': books})


def book_details(request, pk):
    book = Book.objects.get(pk=pk)
    form = CommentForm()

    return render(request, "books/details.html", {'book':book, 'form':form})

# POST books/1/comment
@require_http_methods(["POST"])
def book_comment(request, pk):
    form = CommentForm(data=request.POST)
    if form.is_valid():
        comment = form.save(commit=False)
        book=Book.objects.get(pk=pk)
        comment.book = book
        comment.save()
        messages.add_message(request, messages.INFO, 'Comment is added')
        return HttpResponseRedirect(reverse('books:details', args=(pk,)))
    else:
        messages.add_message(request, messages.WARNING, form.errors)

    return HttpResponseRedirect(reverse('books:details', args=(pk,)))

def books_to_excel(request):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = "attachment; filename=List_{date}.xlsx".format(
        date=timezone.now().strftime("%Y-%m-%d")
    )
    workbook = Workbook()
    ws = workbook.active

    books = Book.objects.all()
    page = request.GET.get('page')
    paginator = Paginator(books, 10)
    books = paginator.get_page(page)

    font = Font(name='Calibri', size = 14, bold = True, italic = False, vertAlign = None, underline = 'none', strike = False, color = 'FF000000')

    ws.append(['tytul', 'strony', 'cena', 'ocena'])
    for b in books:
        ws.append([b.title, b.pages, b.price, b.rating])

    ws['A1'].font = font
    ws['B1'].font = font
    ws['C1'].font = font
    ws['D1'].font = font
    workbook.save(response)
    return response